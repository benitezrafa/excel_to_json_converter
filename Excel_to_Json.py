import json
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def excel_to_dict(excel_file):
    workbook = load_workbook(filename=excel_file, read_only=True)
    result = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        result[sheet_name] = []
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: value for i, value in enumerate(row) if headers[i] is not None}
            result[sheet_name].append(row_dict)
    return result

def convert_excel_to_json(excel_file, json_file):
    try:
        data = excel_to_dict(excel_file)
        
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        print(f"Conversión exitosa. El archivo JSON se ha guardado como '{json_file}'.")
    except Exception as e:
        print(f"Error durante la conversión: {str(e)}")

def select_excel_file():
    root = tk.Tk()
    root.withdraw()  # Oculta la ventana principal de tkinter
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    return file_path

if __name__ == "__main__":
    print("Por favor, selecciona el archivo Excel (.xlsx) en la ventana emergente.")
    excel_file = select_excel_file()
    
    if excel_file:
        json_file = excel_file.rsplit('.', 1)[0] + '.json'
        convert_excel_to_json(excel_file, json_file)
    else:
        print("No se seleccionó ningún archivo.")